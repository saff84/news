"""Extract structured table data from PDF, images (screenshots), and Excel."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

import pdfplumber
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract


@dataclass
class ParsedCell:
    """Single cell: value and optional change_pct in parentheses."""
    value: float | None
    change_pct: float | None
    raw: str


@dataclass
class ParsedRow:
    """Row with indicator name and values per period."""
    indicator_name: str
    unit: str | None
    values: dict[str, ParsedCell]  # period -> cell
    level: int = 0  # 0=main, 1=sub-indicator (— МКД, — ИЖС)


def _parse_number_with_pct(s: str) -> ParsedCell:
    """
    Parse cell value with optional change.
    Supports: '96,4' | '41,6 (+55,1%)' | '7,49 (-0,19 п.п.)' | multiline '96,4\\n(+55,1%)'
    """
    s = (s or "").replace("\n", " ").replace("\r", " ").strip()
    if not s or s == "—" or s == "-":
        return ParsedCell(value=None, change_pct=None, raw=s)

    value = None
    change_pct = None

    # Primary value: decimal (96,4 or 7.49) or integer. Prefer decimal to avoid matching years.
    num_match = re.search(r"(\d{1,3}[,.]\d{1,6})|(\d{1,3}(?:[,.]\d+)?)", s)
    if num_match:
        num_str = (num_match.group(1) or num_match.group(2) or "").replace(",", ".").replace(" ", "")
        try:
            val = float(num_str)
            # Reject 4-digit integers 1900-2100 (likely years in period headers)
            if 1900 <= val <= 2100 and val == int(val) and "." not in (num_match.group(0) or ""):
                pass
            else:
                value = val
        except (ValueError, TypeError):
            pass

    # Change: (+55,1%) or (-0,19 п.п.) or (+1,02 п.п.)
    change_match = re.search(r"\(([+-]?[\d,\.]+)\s*(?:%|п\.п\.?)\)", s, re.I)
    if change_match:
        try:
            change_pct = float(change_match.group(1).replace(",", "."))
        except ValueError:
            pass

    return ParsedCell(value=value, change_pct=change_pct, raw=s)


# Russian month names for date detection
_RU_MONTHS = r"(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)"
# Pattern: "16 февраля 2026 г." or "22 декабря 2025 г. - 15 февраля 2026 г."
_DATE_RANGE_RE = re.compile(
    r"(\d{1,2}\s+" + _RU_MONTHS + r"\s+\d{4}\s*г\.(?:\s*-\s*\d{1,2}\s+" + _RU_MONTHS + r"\s+\d{4}\s*г\.)?)",
    re.I,
)
# Rate / percent values: 15,50 | 16.00 | 7,5 (avoid 4-digit years)
_RATE_VALUE_RE = re.compile(r"(?<!\d)(\d{1,2}[,.]\d{1,2}|\d{1,2})(?!\d)")

_RU_MONTH_NAMES_GENITIVE = (
    "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
)


def _parse_rate_value(num_str: str) -> float | None:
    s = (num_str or "").replace(",", ".").strip()
    if not s:
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    if 1900 <= val <= 2100 and val == int(val):
        return None
    if val < 0 or val > 100:
        return None
    return val


def _rate_value_in_text(s: str) -> float | None:
    for m in _RATE_VALUE_RE.finditer(s or ""):
        val = _parse_rate_value(m.group(1))
        if val is not None:
            return val
    return None


def _dedupe_flat_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[tuple[str, str]] = set()
    out: list[dict[str, Any]] = []
    for r in rows:
        key = (str(r.get("indicator_name") or ""), str(r.get("period") or ""))
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def _format_excel_period(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        d = cell.date()
    elif isinstance(cell, date):
        d = cell
    else:
        s = str(cell).strip()
        return s
    month = _RU_MONTH_NAMES_GENITIVE[d.month - 1]
    return f"{d.day} {month} {d.year} г."


def _cell_to_str(cell: Any) -> str:
    if cell is None:
        return ""
    if isinstance(cell, (datetime, date)):
        return _format_excel_period(cell)
    if isinstance(cell, float) and cell > 30000:
        try:
            from openpyxl.utils.datetime import from_excel

            return _format_excel_period(from_excel(cell))
        except Exception:
            pass
    return str(cell).strip()


def _extract_date_value_pairs_from_text(
    text: str,
    default_indicator: str = "Ключевая ставка",
) -> list[dict[str, Any]]:
    """
    Find all date/range + rate pairs in text (handles OCR merging rows into one line).
    """
    flat: list[dict[str, Any]] = []
    normalized = re.sub(r"\s+", " ", (text or "").replace("\n", " "))
    for m in _DATE_RANGE_RE.finditer(normalized):
        period = m.group(1).strip()
        tail = normalized[m.end() : m.end() + 80]
        value = _rate_value_in_text(tail)
        if value is None:
            continue
        flat.append({
            "indicator_name": default_indicator,
            "period": period,
            "value": value,
            "change_pct": None,
            "unit": "%",
        })
    return flat


def _extract_date_value_table(text: str, default_indicator: str = "Ключевая ставка") -> list[dict[str, Any]]:
    """
    Extract rows from format: "период | значение | ссылка" (e.g. CBR key rate table).
    Each line: date/range, number, optional reference.
    """
    flat: list[dict[str, Any]] = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    for line in lines:
        dm = _DATE_RANGE_RE.search(line)
        if not dm:
            continue
        period = dm.group(1).strip()
        value = _rate_value_in_text(line[dm.end() :])
        if value is None:
            parts = re.split(r"\s{2,}|\t|\|", line)
            if len(parts) >= 2:
                value = _rate_value_in_text(parts[1])
        if value is None:
            continue

        flat.append({
            "indicator_name": default_indicator,
            "period": period,
            "value": value,
            "change_pct": None,
            "unit": "%",
        })

    pairs = _extract_date_value_pairs_from_text(text, default_indicator=default_indicator)
    return _dedupe_flat_rows(flat + pairs)


def _extract_table_from_text(text: str, period_headers: list[str] | None = None) -> list[ParsedRow]:
    """
    Heuristic extraction from OCR/plain text.
    period_headers: e.g. ['2021','2022','2023','2024','2025','Янв. 2026']
    """
    rows: list[ParsedRow] = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    if not period_headers:
        # Try to detect headers from first lines
        period_headers = []
        for line in lines[:5]:
            # Look for year-like tokens
            tokens = re.findall(r"\b(20\d{2}|Янв\.\s*20\d{2}|Фев\.\s*20\d{2})\b", line, re.I)
            if tokens:
                period_headers = tokens
                break

    for line in lines:
        # Skip header-like lines
        if re.match(r"^№?\s*$", line) or re.match(r"^Наименование", line, re.I):
            continue
        # Split by multiple spaces or tabs
        parts = re.split(r"\s{2,}|\t", line)
        if len(parts) < 2:
            continue
        name = parts[0].strip()
        if not name or re.match(r"^\d+$", name):
            continue
        values = {}
        for i, p in enumerate(parts[1:]):
            period = period_headers[i] if period_headers and i < len(period_headers) else str(i)
            values[period] = _parse_number_with_pct(p)
        rows.append(ParsedRow(indicator_name=name, unit=None, values=values, level=0))

    return rows


def _to_flat_rows(rows: list[dict]) -> list[dict[str, Any]]:
    """Convert nested {indicator_name, values: {period: cell}} to flat [{indicator_name, period, value, change_pct}]."""
    flat: list[dict[str, Any]] = []
    for r in rows:
        name = r.get("indicator_name") or ""
        unit = r.get("unit")
        for period, cell in (r.get("values") or {}).items():
            if isinstance(cell, dict):
                v = cell.get("value")
                if v is not None:
                    flat.append({
                        "indicator_name": name,
                        "period": str(period),
                        "value": float(v),
                        "change_pct": cell.get("change_pct"),
                        "unit": unit,
                    })
    return flat


def _looks_like_date_period(s: str) -> bool:
    """Check if string looks like a date/period (e.g. '16 февраля 2026 г.' or '24 июля 2022 г. - 17 декабря 2023 г.')."""
    if not s or len(s) < 10:
        return False
    return bool(_DATE_RANGE_RE.search(s))


def _extract_date_value_from_table_cells(table: list[list[str | None]]) -> list[dict[str, Any]]:
    """Extract from CBR-style table: period | value | reference. Period must look like a date."""
    flat: list[dict[str, Any]] = []
    for row in table or []:
        if not row:
            continue
        # First cell: period (date or range)
        period = (row[0] or "").strip()
        if not period or "Наименование" in period or period == "№":
            continue
        if not _looks_like_date_period(period):
            continue
        value = None
        # Find number in remaining cells
        for cell in row[1:]:
            if cell is None:
                continue
            s = str(cell).strip()
            value = _rate_value_in_text(s)
            if value is not None:
                break
        if value is not None and period:
            flat.append({
                "indicator_name": "Ключевая ставка",
                "period": period,
                "value": value,
                "change_pct": None,
                "unit": "%",
            })
    return flat


def _is_period_header(cell: str) -> bool:
    """Check if cell looks like a period (year or month.year)."""
    s = (cell or "").strip()
    if re.match(r"^20\d{2}$", s):  # 2021, 2022...
        return True
    if re.match(r"^(?:Янв|Фев|Мар|Апр|Май|Июн|Июл|Авг|Сен|Окт|Ноя|Дек)\.?\s*20\d{2}$", s, re.I):
        return True
    return False


def _extract_matrix_table(table: list[list[str | None]]) -> list[dict[str, Any]]:
    """
    Extract from matrix: rows = indicators, cols = periods (2021, 2022, ..., Янв. 2026).
    First 1-2 columns: № and/or Наименование. Rest: period headers.
    Handles hierarchical rows (— sub-indicator) and cells with value (change).
    """
    nested: list[dict[str, Any]] = []
    if not table or len(table) < 2:
        return []

    # Find header row and period columns
    header_row = table[0]
    name_col = 0
    if header_row and len(header_row) > 1:
        first = (header_row[0] or "").strip()
        second = (header_row[1] or "").strip() if len(header_row) > 1 else ""
        if first == "№" and second and ("Наименование" in second or "показателя" in second.lower()):
            name_col = 1
    period_cols: list[str] = []
    for i, h in enumerate(header_row or []):
        if i <= name_col:
            continue
        s = str(h or "").strip()
        if s:
            period_cols.append(s)

    if not period_cols:
        return []

    parent_name = ""
    for row in table[1:]:
        if not row:
            continue
        name = (row[name_col] if name_col < len(row) else row[0] if row else "") or ""
        name = str(name).strip()
        if not name or name == "№" or re.match(r"^Наименование", name, re.I):
            continue
        # Category headers (no numbers) - update parent for sub-rows
        values = {}
        for i, cell in enumerate(row[name_col + 1 : name_col + 1 + len(period_cols)]):
            period = period_cols[i] if i < len(period_cols) else str(i)
            if cell is not None and str(cell).strip():
                pc = _parse_number_with_pct(str(cell))
                if pc.value is not None:
                    values[period] = {"value": pc.value, "change_pct": pc.change_pct, "raw": pc.raw}
        if values:
            if name.startswith("—"):
                full_name = f"{parent_name} — {name.lstrip('—').strip()}" if parent_name else name.lstrip("—").strip()
            else:
                parent_name = name
                full_name = name
            nested.append({
                "indicator_name": full_name,
                "unit": None,
                "values": values,
                "level": 1 if name.startswith("—") else 0,
            })

    return _to_flat_rows(nested)


def extract_from_pdf(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract tables from PDF. Returns flat list of dicts for editing/saving.
    Supports: CBR-style (period|value), matrix (indicators x periods), date-value text.
    """
    flat_rows: list[dict[str, Any]] = []
    nested_rows: list[dict[str, Any]] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables or []:
                if not table or len(table) < 2:
                    continue
                # Try CBR-style first (period | value | ref)
                date_val = _extract_date_value_from_table_cells(table)
                if date_val:
                    flat_rows.extend(date_val)
                    continue
                # Matrix table: indicators x periods (e.g. housing sector report)
                matrix_rows = _extract_matrix_table(table)
                if matrix_rows:
                    flat_rows.extend(matrix_rows)
                    continue
                # Fallback: headers + data
                headers = table[0]
                period_cols = [str(h).strip() for h in (headers or [])[1:] if h and str(h).strip()]
                for row in table[1:]:
                    if not row:
                        continue
                    name = (row[0] or "").strip()
                    if not name or name == "№" or "Наименование" in name:
                        continue
                    values = {}
                    for i, cell in enumerate(row[1:]):
                        period = period_cols[i] if i < len(period_cols) else str(i)
                        if cell is not None and str(cell).strip():
                            pc = _parse_number_with_pct(str(cell))
                            if pc.value is not None:
                                values[period] = {"value": pc.value, "change_pct": pc.change_pct, "raw": pc.raw}
                    if values:
                        nested_rows.append({
                            "indicator_name": name,
                            "unit": None,
                            "values": values,
                            "level": 1 if name.startswith("—") else 0,
                        })

    if flat_rows:
        return flat_rows
    if nested_rows:
        return _to_flat_rows(nested_rows)
    # Fallback: extract text and try date-value pattern
    text = ""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"
    return _extract_date_value_table(text)


def _preprocess_for_ocr(img: Image.Image) -> Image.Image:
    """Upscale small screenshots and improve contrast for Tesseract."""
    if img.mode != "RGB":
        img = img.convert("RGB")
    w, h = img.size
    if w < 1400:
        scale = 1400 / max(w, 1)
        img = img.resize((int(w * scale), int(h * scale)), Image.Resampling.LANCZOS)
    gray = img.convert("L")
    gray = ImageEnhance.Contrast(gray).enhance(1.4)
    gray = gray.filter(ImageFilter.SHARPEN)
    return gray


def _ocr_collect_text(img: Image.Image) -> str:
    """Run OCR with several page layouts; merge text for downstream parsers."""
    pre = _preprocess_for_ocr(img)
    chunks: list[str] = []
    for psm in (6, 4, 11):
        try:
            chunks.append(
                pytesseract.image_to_string(pre, lang="rus+eng", config=f"--psm {psm}")
            )
        except Exception:
            pass
    try:
        data = pytesseract.image_to_data(pre, lang="rus+eng", output_type=pytesseract.Output.DICT)
        by_line: dict[tuple[int, int], list[str]] = {}
        for i, word in enumerate(data.get("text") or []):
            w = (word or "").strip()
            if not w:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except (ValueError, TypeError):
                conf = -1
            if conf >= 0 and conf < 40:
                continue
            key = (int(data["block_num"][i]), int(data["line_num"][i]))
            by_line.setdefault(key, []).append(w)
        line_text = "\n".join(" ".join(words) for _, words in sorted(by_line.items()))
        if line_text.strip():
            chunks.append(line_text)
    except Exception:
        pass
    return "\n".join(chunks)


def _extract_key_rate_rows(text: str) -> list[dict[str, Any]]:
    """CBR-style period + rate; combine line-based and global pair extraction."""
    line_rows = _extract_date_value_table(text)
    if len(line_rows) >= 2:
        return line_rows
    pairs = _extract_date_value_pairs_from_text(text)
    merged = _dedupe_flat_rows(line_rows + pairs)
    return merged if merged else line_rows


def extract_from_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    """Extract tables from the first worksheet (.xlsx)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        flat: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            table: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                if row is None or all(c is None or str(c).strip() == "" for c in row):
                    continue
                table.append([_cell_to_str(c) for c in row])
            if len(table) < 2:
                continue
            date_val = _extract_date_value_from_table_cells(table)
            if date_val:
                flat.extend(date_val)
                continue
            matrix = _extract_matrix_table(table)
            if matrix:
                flat.extend(matrix)
                continue
            for data_row in table[1:]:
                if not data_row:
                    continue
                period = (data_row[0] or "").strip()
                if not _looks_like_date_period(period):
                    continue
                value = None
                for cell in data_row[1:]:
                    value = _rate_value_in_text(str(cell or ""))
                    if value is not None:
                        break
                if value is not None:
                    flat.append({
                        "indicator_name": "Ключевая ставка",
                        "period": period,
                        "value": value,
                        "change_pct": None,
                        "unit": "%",
                    })
        return _dedupe_flat_rows(flat)
    finally:
        wb.close()


def extract_from_image(file_bytes: bytes) -> list[dict[str, Any]]:
    """
    Extract text from image via OCR, then parse table heuristically.
    """
    img = Image.open(io.BytesIO(file_bytes))
    text = _ocr_collect_text(img)

    key_rate = _extract_key_rate_rows(text)
    if key_rate:
        return key_rate

    rows = _extract_table_from_text(text)
    nested = [
        {
            "indicator_name": r.indicator_name,
            "unit": r.unit,
            "values": {
                k: {"value": v.value, "change_pct": v.change_pct, "raw": v.raw}
                for k, v in r.values.items()
                if v.value is not None
            },
            "level": r.level,
        }
        for r in rows
    ]
    flat = _to_flat_rows(nested)
    return flat if flat else key_rate
